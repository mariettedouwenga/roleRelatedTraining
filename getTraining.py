import openpyxl
import pandas as pd
import streamlit as st
import markdown

header = st.container()

# load excel with its path 
trainingMatrix = openpyxl.load_workbook("TrainingMatrix.xlsx") 


trainingMatrix_sh = trainingMatrix.active

with header:
    st.title('Infor LN Role Related Training')
    st.image("PR.jpg")
    st.divider()
    st.subheader('Find the training related to your Infor LN role.')
    roleData = pd.read_csv('inforRoles.csv')


    role = st.selectbox('Select a role', roleData, index = None,
    placeholder="Select your Infor LN role",)
      
# iterate throught columns in excel sheet and display data 
    for j in range(1, trainingMatrix_sh.max_column+1): 

        #Iterate through rows in excel sheet
        for i in range(1, trainingMatrix_sh.max_row+1): 
            cell_obj = trainingMatrix_sh.cell(i, column=j)
            role_name = trainingMatrix_sh.cell(1,column=j)
            training_catagory = trainingMatrix_sh.cell(i,column=1)
            training_name = trainingMatrix_sh.cell(row=i,column=2)
            training_description = trainingMatrix_sh.cell(row=i,column=3)    
                 
            
            if cell_obj.value == "x":
                if role_name.value == role:                                                         
                                                
                    st.write(f"**{training_name.value}**")
                    st.text("Description: " + "\n" + training_description.value)
                    st.link_button("Find training dates", "https://boskalis.sharepoint.com/sites/D001418/_layouts/15/Events.aspx?Page=%2Fsites%2FD001418%2FSitePages%2FTraining-Dates.aspx&InstanceId=5fced5f5-7ebc-42a2-87bf-cdc70e4868ea&Category=Training&StartDate=2025-03-07&EndDate=2025-12-31&AudienceTarget=false")
                    st.divider()
    

               

    

                